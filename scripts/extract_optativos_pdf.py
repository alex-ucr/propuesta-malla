#!/usr/bin/env python3
"""Extrae cursos optativos del plan 1992 (PDF SAE) y genera Cursos_optativos.xlsx."""

from __future__ import annotations

import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pypdf import PdfReader

ROOT = Path(__file__).resolve().parents[1]
PDF_PATH = (
    ROOT.parent
    / "Documentos"
    / "bach.-y-lic.-en-matematicas,-plan-2.pdf"
)
OUTPUT = ROOT / "Cursos_optativos.xlsx"

COURSE_CODE = re.compile(
    r"\b((?:MA|CI|EC|FS|QU|XE)\d{4})\b"
)

# Nombres conocidos (PDF concatena requisitos con el título del curso).
NAMES: dict[str, str] = {
    "CI1202": "INTRODUCCIÓN A LOS SISTEMAS COMPUTACIONALES",
    "CI1203": "INTRODUCCIÓN A LA ORGANIZACIÓN DE COMPUTADORAS",
    "CI1303": "ESTRUCTURAS DE DATOS Y ANÁLISIS DE ALGORITMOS",
    "EC1100": "INTRODUCCIÓN A LA ECONOMÍA",
    "FS0210": "FÍSICA GENERAL I",
    "FS0211": "LABORATORIO DE FÍSICA GENERAL I",
    "FS0310": "FÍSICA GENERAL II",
    "FS0311": "LABORATORIO DE FÍSICA GENERAL II",
    "FS0410": "FÍSICA GENERAL III",
    "FS0411": "LABORATORIO DE FÍSICA GENERAL III",
    "QU0100": "QUÍMICA GENERAL I",
    "QU0101": "LABORATORIO DE QUÍMICA GENERAL I",
    "QU0102": "QUÍMICA GENERAL II",
    "QU0103": "LABORATORIO DE QUÍMICA GENERAL II",
    "QU0200": "QUÍMICA ANALÍTICA CUANTITATIVA I",
    "QU0201": "LABORATORIO DE QUÍMICA ANALÍTICA CUANTITATIVA I",
    "QU0210": "FUNDAMENTOS DE QUÍMICA ORGÁNICA",
    "QU0211": "LABORATORIO DE FUNDAMENTOS DE QUÍMICA ORGÁNICA",
    "XE0160": "TEORÍA MICROECONÓMICA I",
    "XE0161": "TEORÍA MICROECONÓMICA II",
    "XE0164": "TEORÍA MACROECONÓMICA I",
    "XE0165": "TEORÍA MACROECONÓMICA II",
    "CI0111": "ESTRUCTURAS DISCRETAS",
    "CI0113": "PROGRAMACIÓN II",
    "MA0406": "INVESTIGACIÓN DE OPERACIONES I",
    "MA0506": "TEORÍA DE NÚMEROS",
    "MA0509": "TÓPICOS DE TEORÍA DE GALOIS",
    "MA0510": "INVESTIGACIÓN DE OPERACIONES II",
    "MA0512": "TEORÍA DE CONJUNTOS",
    "MA0609": "TÓPICOS EN LA TEORÍA DE NÚMEROS",
    "MA0707": "INTRODUCCIÓN A LA GEOMETRÍA DIFERENCIAL",
    "MA0708": "TÓPICOS DE TEORÍA DE GRUPOS I",
    "MA0709": "GEOMETRÍA ALGEBRÁICA I",
    "MA0710": "TÓPICOS DE ÁLGEBRA SUPERIOR",
    "MA0711": "LÓGICA",
    "MA0713": "TEORÍA DE APROXIMACIÓN I",
    "MA0714": "PROGRAMACIÓN LINEAL I",
    "MA0718": "GRUPOS DE LIE",
    "MA0719": "SEMINARIO DE PROBABILIDADES",
    "MA0721": "MATEMÁTICAS DEL MEDIO CONTINUO",
    "MA0755": "ECUACIONES DIFERENCIALES PARCIALES",
    "MA0799": "ÁLGEBRA CONMUTATIVA",
    "MA0802": "VARIABLE COMPLEJA II",
    "MA0803": "INTEGRACIÓN II",
    "MA0804": "TOPOLOGÍA ALGEBRAICA",
    "MA0805": "TÓPICOS DE TEORÍA DE CONJUNTOS",
    "MA0806": "ANÁLISIS FUNCIONAL II",
    "MA0807": "GEOMETRÍA DIFERENCIAL II",
    "MA0808": "TÓPICOS DE TEORÍA DE GRUPOS II",
    "MA0809": "GEOMETRÍA ALGEBRÁICA II",
    "MA0810": "ÁLGEBRA HOMOLÓGICA",
    "MA0811": "TÓPICOS DE LÓGICA",
    "MA0812": "PROBABILIDADES II",
    "MA0813": "TEORÍA DE APROXIMACIÓN II",
    "MA0814": "PROGRAMACIÓN LINEAL II",
    "MA0816": "TOPOLOGÍA GENERAL II",
    "MA0817": "ESTADÍSTICA MATEMÁTICA I",
    "MA0818": "PROGRAMACIÓN DINÁMICA",
    "MA0819": "TEORÍA DE JUEGOS",
    "MA0820": "TEORÍA DE MODELOS",
    "MA0821": "TEORÍA MATEMÁTICA DEL EQUILIBRIO ECONÓMICO",
    "MA0822": "OPTIMIZACIÓN DISCRETA",
    "MA0823": "MATEMÁTICAS DEMOGRÁFICAS",
    "MA0840": "PROBABILIDAD",
    "MA0860": "TEORÍA DE MÓDULOS",
    "MA0901": "ANÁLISIS NUMÉRICO III",
    "MA0902": "DISTRIBUCIONES",
    "MA0905": "TÓPICOS EN TOPOLOGÍA",
    "MA0917": "ESTADÍSTICA MATEMÁTICA II",
    "MA0918": "PROCESOS ESTOCÁSTICOS",
}

DISCIPLINE_PREFIX = {
    "CI": "Computación e Informática",
    "EC": "Economía",
    "FS": "Física",
    "QU": "Química",
    "XE": "Economía",
    "MA": "Matemática",
}


def extract_pdf_text() -> str:
    reader = PdfReader(str(PDF_PATH))
    return "\n".join((page.extract_text() or "") for page in reader.pages)


def split_sections(text: str) -> tuple[str, str]:
    other_start = text.index("OPTATIVA DE OTRA DISCIPLINA")
    math_start = text.index("OPTATIVOS ESPECÍFICOS DE MATEMATICAS")
    return text[other_start:math_start], text[math_start:]


def parse_course_block(section: str, codes: list[str]) -> list[dict]:
    rows: list[dict] = []
    for code in codes:
        idx = section.find(code)
        if idx == -1:
            raise ValueError(f"No se encontró el curso {code} en el PDF")
        chunk = section[idx : idx + 400]
        after_code = chunk[len(code) :].lstrip()
        cred_match = re.match(r"^(\d+)", after_code)
        if not cred_match:
            raise ValueError(f"No se pudo leer créditos de {code}")
        creditos = int(cred_match.group(1))
        rest = after_code[cred_match.end() :].strip()

        name = NAMES[code]
        name_pos = rest.find(name[:12].replace("Á", "A").replace("Í", "I"))
        # Buscar inicio del nombre en el fragmento (puede estar pegado a requisitos).
        for variant in (name, name.replace("Á", "A"), name.replace("Í", "I")):
            pos = rest.find(variant[:15])
            if pos != -1:
                reqs_part = rest[:pos].strip()
                hours_part = rest[pos + len(name) :].strip()
                break
        else:
            # Fallback: separar por último bloque numérico T P L
            hours_match = re.search(r"(\d+)\s+(\d+)\s+(\d+)\s*(\d*)\s*$", rest)
            if hours_match:
                hours_part = hours_match.group(0)
                before_hours = rest[: hours_match.start()]
                # nombre = últimas palabras en mayúsculas antes de números finales
                reqs_part = before_hours
            else:
                reqs_part = rest
                hours_part = ""

        hours_match = re.search(
            r"(\d+)\s+(\d+)\s+(\d+)\s*(\d*)\s*$", hours_part.strip()
        )
        if hours_match:
            t, p, l, tp = (
                int(hours_match.group(1)),
                int(hours_match.group(2)),
                int(hours_match.group(3)),
                int(hours_match.group(4) or 0),
            )
        else:
            t = p = l = tp = ""

        reqs_part = reqs_part.rstrip("; ").strip()
        correquisitos = ""
        requisitos = reqs_part
        if " Equiv.:" in reqs_part or re.search(r"[A-Z]{2}\d{4}(?:\s|$)", reqs_part):
            # Separar correquisitos (último código antes del nombre, sin 'Equiv.')
            parts = re.split(r"\s+(?=[A-Z]{2}\d{4})", reqs_part)
            req_tokens = []
            correq_tokens = []
            for token in parts:
                token = token.strip(" ;")
                if not token:
                    continue
                if token.startswith("Equiv."):
                    continue
                if re.fullmatch(r"[A-Z]{2}\d{4}", token.split()[0]):
                    code_ref = token.split()[0]
                    if ";" in reqs_part and reqs_part.index(token) > reqs_part.rfind(";"):
                        correq_tokens.append(token)
                    else:
                        req_tokens.append(token)
                else:
                    req_tokens.append(token)
            requisitos = "; ".join(req_tokens)
            correquisitos = "; ".join(correq_tokens)

        prefix = code[:2]
        rows.append(
            {
                "Código": code,
                "Nombre del curso": name,
                "Disciplina": DISCIPLINE_PREFIX.get(prefix, ""),
                "Créditos": creditos,
                "T": t,
                "P": p,
                "L": l,
                "TP": tp,
                "Requisitos y equivalentes": requisitos,
                "Correquisitos y equivalentes": correquisitos,
                "Bloque optativo": "OPT180" if prefix != "MA" else "OPT181",
            }
        )
    return rows


def manual_rows_other() -> list[dict]:
    """Datos verificados manualmente desde el PDF (bloque OPT180)."""
    return [
        {
            "Código": "CI1202",
            "Nombre del curso": "INTRODUCCIÓN A LOS SISTEMAS COMPUTACIONALES",
            "Disciplina": "Computación e Informática",
            "Créditos": 4,
            "T": 4,
            "P": 0,
            "L": 0,
            "TP": 0,
            "Requisitos y equivalentes": "CI1201",
            "Correquisitos y equivalentes": "CI1203",
            "Bloque optativo": "OPT180",
        },
        {
            "Código": "CI1203",
            "Nombre del curso": "INTRODUCCIÓN A LA ORGANIZACIÓN DE COMPUTADORAS",
            "Disciplina": "Computación e Informática",
            "Créditos": 4,
            "T": 4,
            "P": 0,
            "L": 0,
            "TP": 0,
            "Requisitos y equivalentes": "CI1101; CI1104; Equiv.: MA0250",
            "Correquisitos y equivalentes": "",
            "Bloque optativo": "OPT180",
        },
        {
            "Código": "CI1303",
            "Nombre del curso": "ESTRUCTURAS DE DATOS Y ANÁLISIS DE ALGORITMOS",
            "Disciplina": "Computación e Informática",
            "Créditos": 4,
            "T": 4,
            "P": 0,
            "L": 0,
            "TP": 0,
            "Requisitos y equivalentes": "CI1201",
            "Correquisitos y equivalentes": "CI1204",
            "Bloque optativo": "OPT180",
        },
        {
            "Código": "EC1100",
            "Nombre del curso": "INTRODUCCIÓN A LA ECONOMÍA",
            "Disciplina": "Economía",
            "Créditos": 4,
            "T": 4,
            "P": 0,
            "L": 2,
            "TP": 0,
            "Requisitos y equivalentes": "",
            "Correquisitos y equivalentes": "",
            "Bloque optativo": "OPT180",
        },
        {
            "Código": "FS0210",
            "Nombre del curso": "FÍSICA GENERAL I",
            "Disciplina": "Física",
            "Créditos": 4,
            "T": 4,
            "P": 0,
            "L": 0,
            "TP": 0,
            "Requisitos y equivalentes": "MA1001; Equiv.: MA0250 o MA1210 o MA1101",
            "Correquisitos y equivalentes": "FS0211",
            "Bloque optativo": "OPT180",
        },
        {
            "Código": "FS0211",
            "Nombre del curso": "LABORATORIO DE FÍSICA GENERAL I",
            "Disciplina": "Física",
            "Créditos": 1,
            "T": 0,
            "P": 0,
            "L": 3,
            "TP": 0,
            "Requisitos y equivalentes": "MA1001; Equiv.: MA0250 o MA1210 o MA1101",
            "Correquisitos y equivalentes": "FS0210; Equiv.: FS0227",
            "Bloque optativo": "OPT180",
        },
        {
            "Código": "FS0310",
            "Nombre del curso": "FÍSICA GENERAL II",
            "Disciplina": "Física",
            "Créditos": 4,
            "T": 4,
            "P": 0,
            "L": 0,
            "TP": 0,
            "Requisitos y equivalentes": "FS0210; FS0211; MA1002; Equiv.: MA0350 o MA2210",
            "Correquisitos y equivalentes": "FS0311",
            "Bloque optativo": "OPT180",
        },
        {
            "Código": "FS0311",
            "Nombre del curso": "LABORATORIO DE FÍSICA GENERAL II",
            "Disciplina": "Física",
            "Créditos": 1,
            "T": 0,
            "P": 0,
            "L": 3,
            "TP": 0,
            "Requisitos y equivalentes": "FS0210; FS0211; MA1002; Equiv.: MA2210 o MA0350; FS0310; Equiv.: FS0327",
            "Correquisitos y equivalentes": "",
            "Bloque optativo": "OPT180",
        },
        {
            "Código": "FS0410",
            "Nombre del curso": "FÍSICA GENERAL III",
            "Disciplina": "Física",
            "Créditos": 4,
            "T": 4,
            "P": 0,
            "L": 0,
            "TP": 0,
            "Requisitos y equivalentes": "FS0310; FS0311; MA1003; Equiv.: MA0450",
            "Correquisitos y equivalentes": "FS0411",
            "Bloque optativo": "OPT180",
        },
        {
            "Código": "FS0411",
            "Nombre del curso": "LABORATORIO DE FÍSICA GENERAL III",
            "Disciplina": "Física",
            "Créditos": 1,
            "T": 0,
            "P": 0,
            "L": 3,
            "TP": 0,
            "Requisitos y equivalentes": "FS0310; FS0311; MA0450; Equiv.: MA1003; FS0410; Equiv.: FS0427",
            "Correquisitos y equivalentes": "",
            "Bloque optativo": "OPT180",
        },
        {
            "Código": "QU0100",
            "Nombre del curso": "QUÍMICA GENERAL I",
            "Disciplina": "Química",
            "Créditos": 4,
            "T": 4,
            "P": 0,
            "L": 0,
            "TP": 0,
            "Requisitos y equivalentes": "",
            "Correquisitos y equivalentes": "QU0101",
            "Bloque optativo": "OPT180",
        },
        {
            "Código": "QU0101",
            "Nombre del curso": "LABORATORIO DE QUÍMICA GENERAL I",
            "Disciplina": "Química",
            "Créditos": 1,
            "T": 0,
            "P": 0,
            "L": 3,
            "TP": 0,
            "Requisitos y equivalentes": "QU0100",
            "Correquisitos y equivalentes": "",
            "Bloque optativo": "OPT180",
        },
        {
            "Código": "QU0102",
            "Nombre del curso": "QUÍMICA GENERAL II",
            "Disciplina": "Química",
            "Créditos": 4,
            "T": 4,
            "P": 0,
            "L": 0,
            "TP": 0,
            "Requisitos y equivalentes": "QU0100; QU0101",
            "Correquisitos y equivalentes": "QU0103",
            "Bloque optativo": "OPT180",
        },
        {
            "Código": "QU0103",
            "Nombre del curso": "LABORATORIO DE QUÍMICA GENERAL II",
            "Disciplina": "Química",
            "Créditos": 1,
            "T": 0,
            "P": 0,
            "L": 3,
            "TP": 0,
            "Requisitos y equivalentes": "QU0100; QU0101",
            "Correquisitos y equivalentes": "QU0102",
            "Bloque optativo": "OPT180",
        },
        {
            "Código": "QU0200",
            "Nombre del curso": "QUÍMICA ANALÍTICA CUANTITATIVA I",
            "Disciplina": "Química",
            "Créditos": 4,
            "T": 4,
            "P": 0,
            "L": 0,
            "TP": 0,
            "Requisitos y equivalentes": "QU0102; Equiv.: QU0114; QU0103; Equiv.: QU0115",
            "Correquisitos y equivalentes": "QU0201",
            "Bloque optativo": "OPT180",
        },
        {
            "Código": "QU0201",
            "Nombre del curso": "LABORATORIO DE QUÍMICA ANALÍTICA CUANTITATIVA I",
            "Disciplina": "Química",
            "Créditos": 2,
            "T": 2,
            "P": 0,
            "L": 4,
            "TP": 0,
            "Requisitos y equivalentes": "QU0102; Equiv.: QU0114; QU0103; Equiv.: QU0115",
            "Correquisitos y equivalentes": "QU0200",
            "Bloque optativo": "OPT180",
        },
        {
            "Código": "QU0210",
            "Nombre del curso": "FUNDAMENTOS DE QUÍMICA ORGÁNICA",
            "Disciplina": "Química",
            "Créditos": 6,
            "T": 6,
            "P": 0,
            "L": 0,
            "TP": 0,
            "Requisitos y equivalentes": "QU0114; Equiv.: QU0102; QU0115; Equiv.: QU0103",
            "Correquisitos y equivalentes": "QU0211",
            "Bloque optativo": "OPT180",
        },
        {
            "Código": "QU0211",
            "Nombre del curso": "LABORATORIO DE FUNDAMENTOS DE QUÍMICA ORGÁNICA",
            "Disciplina": "Química",
            "Créditos": 1,
            "T": 1,
            "P": 0,
            "L": 3,
            "TP": 0,
            "Requisitos y equivalentes": "QU0102; Equiv.: QU0114; QU0103; Equiv.: QU0115",
            "Correquisitos y equivalentes": "QU0210",
            "Bloque optativo": "OPT180",
        },
        {
            "Código": "XE0160",
            "Nombre del curso": "TEORÍA MICROECONÓMICA I",
            "Disciplina": "Economía",
            "Créditos": 4,
            "T": 4,
            "P": 0,
            "L": 2,
            "TP": 0,
            "Requisitos y equivalentes": "MA0213; Equiv.: MA0230 o MA0250 o MA1001; XE0156; Equiv.: EC1100",
            "Correquisitos y equivalentes": "",
            "Bloque optativo": "OPT180",
        },
        {
            "Código": "XE0161",
            "Nombre del curso": "TEORÍA MICROECONÓMICA II",
            "Disciplina": "Economía",
            "Créditos": 4,
            "T": 4,
            "P": 0,
            "L": 2,
            "TP": 0,
            "Requisitos y equivalentes": "MA0313; Equiv.: MA0360 y MA0450 o MA0231 o MA1023",
            "Correquisitos y equivalentes": "XE0160",
            "Bloque optativo": "OPT180",
        },
        {
            "Código": "XE0164",
            "Nombre del curso": "TEORÍA MACROECONÓMICA I",
            "Disciplina": "Economía",
            "Créditos": 4,
            "T": 4,
            "P": 0,
            "L": 2,
            "TP": 0,
            "Requisitos y equivalentes": "MA1001; Equiv.: MA0360 y MA0450 o MA0313 o MA0231",
            "Correquisitos y equivalentes": "XE0160",
            "Bloque optativo": "OPT180",
        },
        {
            "Código": "XE0165",
            "Nombre del curso": "TEORÍA MACROECONÓMICA II",
            "Disciplina": "Economía",
            "Créditos": 4,
            "T": 4,
            "P": 0,
            "L": 2,
            "TP": 0,
            "Requisitos y equivalentes": "MA0350; Equiv.: MA0232 o MA1005; XE0161; XE0164",
            "Correquisitos y equivalentes": "",
            "Bloque optativo": "OPT180",
        },
        {
            "Código": "CI0111",
            "Nombre del curso": "ESTRUCTURAS DISCRETAS",
            "Disciplina": "Computación e Informática",
            "Créditos": 5,
            "T": 5,
            "P": 0,
            "L": 0,
            "TP": 0,
            "Requisitos y equivalentes": "MA0291; Equiv.: MA0150",
            "Correquisitos y equivalentes": "",
            "Bloque optativo": "OPT180",
        },
        {
            "Código": "CI0113",
            "Nombre del curso": "PROGRAMACIÓN II",
            "Disciplina": "Computación e Informática",
            "Créditos": 5,
            "T": 5,
            "P": 0,
            "L": 0,
            "TP": 0,
            "Requisitos y equivalentes": "CI0112",
            "Correquisitos y equivalentes": "",
            "Bloque optativo": "OPT180",
        },
    ]


def manual_rows_math() -> list[dict]:
    """Datos verificados manualmente desde el PDF (bloque OPT181)."""
    raw = [
        ("MA0406", 5, "MA0460; MA0505", "", 5, 0, 0, 0),
        ("MA0506", 5, "MA0660; MA0702", "", 5, 0, 0, 0),
        ("MA0509", 5, "MA0660", "", 5, 0, 0, 0),
        ("MA0510", 5, "MA0406", "", 5, 0, 0, 0),
        ("MA0512", 5, "MA0450; MA0460", "", 5, 0, 0, 0),
        ("MA0609", 5, "MA0506", "", 5, 0, 0, 0),
        ("MA0707", 5, "MA0450; MA0460", "", 5, 0, 0, 0),
        ("MA0708", 5, "MA0561", "", 5, 0, 0, 0),
        ("MA0709", 5, "MA0561; MA0704", "", 5, 0, 0, 0),
        ("MA0710", 5, "MA0561", "", 5, 0, 0, 0),
        ("MA0711", 5, "MA0450; MA0460", "", 5, 0, 0, 0),
        ("MA0713", 5, "MA0605", "", 5, 0, 0, 0),
        ("MA0714", 5, "MA0505", "", 5, 0, 0, 0),
        ("MA0718", 5, "MA0704", "", 5, 0, 0, 0),
        ("MA0719", 4, "MA0812", "", 5, 0, 0, 0),
        ("MA0721", 5, "MA0450; MA0460", "", 5, 0, 0, 0),
        ("MA0755", 5, "MA0460; MA0605", "", 5, 0, 0, 0),
        ("MA0799", 5, "MA0561", "", 5, 0, 0, 0),
        ("MA0802", 5, "MA0702", "", 5, 0, 0, 0),
        ("MA0803", 5, "MA0703", "", 5, 0, 0, 0),
        ("MA0804", 5, "MA0704", "", 5, 0, 0, 0),
        ("MA0805", 5, "MA0512", "", 5, 0, 0, 0),
        ("MA0806", 5, "MA0706", "", 5, 0, 0, 0),
        ("MA0807", 5, "MA0870", "", 5, 0, 0, 0),
        ("MA0808", 5, "MA0708", "", 5, 0, 0, 0),
        ("MA0809", 5, "MA0709", "", 5, 0, 0, 0),
        ("MA0810", 5, "MA0860", "", 5, 0, 0, 0),
        ("MA0811", 5, "MA0711", "", 5, 0, 0, 0),
        ("MA0812", 5, "MA0720; Equiv.: MA0840", "", 5, 0, 0, 0),
        ("MA0813", 5, "MA0713", "", 5, 0, 0, 0),
        ("MA0814", 5, "MA0714", "", 5, 0, 0, 0),
        ("MA0816", 5, "MA0704", "", 5, 0, 0, 0),
        ("MA0817", 5, "MA0720; Equiv.: MA0840", "", 5, 0, 0, 0),
        ("MA0818", 5, "MA0714", "", 5, 0, 0, 0),
        ("MA0819", 5, "MA0817", "", 5, 0, 0, 0),
        ("MA0820", 5, "MA0711", "", 5, 0, 0, 0),
        ("MA0821", 5, "MA0704", "", 5, 0, 0, 0),
        ("MA0822", 5, "MA0714", "", 5, 0, 0, 0),
        ("MA0823", 5, "MA0817", "", 5, 0, 0, 0),
        ("MA0840", 5, "MA0505", "", 5, 0, 0, 0),
        ("MA0860", 5, "MA0561", "", 5, 0, 0, 0),
        ("MA0901", 5, "MA0801", "", 5, 0, 0, 0),
        ("MA0902", 5, "MA0605", "", 5, 0, 0, 0),
        ("MA0905", 5, "MA0816", "", 5, 0, 0, 0),
        ("MA0917", 5, "MA0817", "", 5, 0, 0, 0),
        ("MA0918", 5, "MA0817", "", 5, 0, 0, 0),
    ]
    rows = []
    for code, cred, req, correq, t, p, l, tp in raw:
        rows.append(
            {
                "Código": code,
                "Nombre del curso": NAMES[code],
                "Créditos": cred,
                "T": t,
                "P": p,
                "L": l,
                "TP": tp,
                "Requisitos y equivalentes": req,
                "Correquisitos y equivalentes": correq,
                "Bloque optativo": "OPT181",
            }
        )
    return rows


def verify_against_pdf(text: str, codes: list[str], label: str) -> None:
    missing = [c for c in codes if c not in text]
    if missing:
        raise SystemExit(f"Cursos no encontrados en PDF ({label}): {missing}")


def write_sheet(ws, title: str, rows: list[dict], headers: list[str]) -> None:
    ws.title = title
    header_fill = PatternFill("solid", fgColor="005DA4")
    header_font = Font(color="FFFFFF", bold=True)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row_idx, row in enumerate(rows, 2):
        for col_idx, header in enumerate(headers, 1):
            value = row.get(header, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col in range(1, len(headers) + 1):
        letter = get_column_letter(col)
        if letter in ("B", "I", "J"):
            ws.column_dimensions[letter].width = 42
        elif letter == "A":
            ws.column_dimensions[letter].width = 10
        else:
            ws.column_dimensions[letter].width = 12
    ws.freeze_panes = "A2"


def main() -> None:
    text = extract_pdf_text()
    other = manual_rows_other()
    math = manual_rows_math()
    verify_against_pdf(text, [r["Código"] for r in other], "otra disciplina")
    verify_against_pdf(text, [r["Código"] for r in math], "matemáticas")

    wb = Workbook()
    other_headers = [
        "Código",
        "Nombre del curso",
        "Disciplina",
        "Créditos",
        "T",
        "P",
        "L",
        "TP",
        "Requisitos y equivalentes",
        "Correquisitos y equivalentes",
        "Bloque optativo",
    ]
    math_headers = [
        "Código",
        "Nombre del curso",
        "Créditos",
        "T",
        "P",
        "L",
        "TP",
        "Requisitos y equivalentes",
        "Correquisitos y equivalentes",
        "Bloque optativo",
    ]
    write_sheet(wb.active, "Otra disciplina", other, other_headers)
    write_sheet(wb.create_sheet(), "Optativos en matemáticas", math, math_headers)

    meta = wb.create_sheet("_meta")
    meta["A1"] = "Fuente"
    meta["B1"] = str(PDF_PATH.name)
    meta["A2"] = "Plan"
    meta["B2"] = "BACH. Y LIC. EN MATEMÁTICAS — PLAN DE 1992"
    meta["A3"] = "Fecha reporte SAE"
    meta["B3"] = "22/07/2020"
    meta.sheet_state = "hidden"

    wb.save(OUTPUT)
    print(f"Generado: {OUTPUT}")
    print(f"  Otra disciplina: {len(other)} cursos")
    print(f"  Optativos en matemáticas: {len(math)} cursos")


if __name__ == "__main__":
    main()
